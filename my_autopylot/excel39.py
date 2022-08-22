import os

output_folder_path = os.path.join(os.path.abspath(
    r'C:\Users\Public\PyBots'), 'My-AutoPylot', 'Excel Folder')

# create output folder if not present
if not os.path.exists(output_folder_path):
    os.makedirs(output_folder_path)


def authenticate_google_spreadsheet(credential_file_path=""):
    """
    Description:
        Authenticates Google Spreadsheet.
    Args:
        credential_file_path (str, optional): Path of credential file. Defaults to "".
    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (object): Google Spreadsheet Auth object.
    """

    # import section
    from my_autopylot.CrashHandler import report_error
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials

    # Response section
    error = None
    status = False
    data = None

    # Logic section
    try:
        if not credential_file_path:
            raise Exception("credential (json) file path cannot be empty")

        scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets',
                 "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]

        creds = ServiceAccountCredentials.from_json_keyfile_name(
            credential_file_path, scope)

        gc = gspread.authorize(creds)

    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
        # # If the function returns a value, it should be assigned to the data variable.
        data = gc

    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]


def excel_get_dataframe_from_google_spreadsheet(auth, spreadsheet_url="", sheet_name="Sheet1"):
    # Description:
    """
    Description:
        Get dataframe from google spreadsheet.
    Args:
        URL (str, optional): (Only in Windows)Name of Window you want to activate.
        Eg: Notepad. Defaults to "".

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (object): Dataframe object.
    """

    # import section
    from my_autopylot.CrashHandler import report_error
    import pandas as pd

    # Response section
    error = None
    status = False
    data = None
    # Logic section

    try:
        if not auth:
            raise Exception(
                "Please call authenticate_google_spreadsheet function to get auth")

        if not spreadsheet_url:
            raise Exception("spreadsheet url cannot be empty")

        sh = auth.open_by_url(url=spreadsheet_url)

        # get all the worksheets from sh
        worksheet_list = sh.worksheets()

        # check if sheet_name is already present in worksheet_list
        sheet_present = False
        for worksheet in worksheet_list:
            if worksheet.title == sheet_name:
                sheet_present = True
                break

        if not sheet_present:
            raise Exception("Sheet name not found")
        else:
            worksheet = sh.worksheet(sheet_name)

        data_frame = pd.DataFrame(worksheet.get_all_records())

    except Exception as ex:
        # check if it is a permission error
        if 'PERMISSION_DENIED' in str(ex):
            raise Exception(
                "Permission Denied. Please share the spreadsheet with client email as per Credential JSON File")
        else:
            report_error(ex)
            error = ex

    else:
        status = True
        data = data_frame
        # # If the function returns a value, it should be assigned to the data variable.

    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]


def excel_tabular_data_from_website(website_url="", table_number=""):
    """
    Description:
        Gets Website Table Data Easily as an Excel using Pandas. Just pass the URL of Website having HTML Tables.
    Args:
        website_url (str, optional): URL of Website. Defaults to "".
        table_number (int, optional): Table Number. Defaults to all.

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (object): Dataframe object.
    """

    # Import Section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response Section
    error = None
    status = False
    data = None

    # Logic Section
    try:
        if not website_url:
            raise Exception("Website URL cannot be empty")

        all_tables = pd.read_html(website_url)

        if not table_number:
            data = all_tables
        else:
            if table_number > len(all_tables):
                raise Exception(
                    "Table number cannot be greater than number of tables")

            if table_number < 1:
                raise Exception("Table number cannot be less than 1")

            data = all_tables[table_number - 1]

    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]


# [status, data]
# browser_get_html_tabular_data_from_website(website_url="https://en.wikipedia.org/wiki/List_of_footballers_with_500_or_more_goals",output_folder=r"C:\Users\mrmay\OneDrive\Desktop\Misc")
# print(excel_tabular_data_from_website(
#     "https://en.wikipedia.org/wiki/Scoring_in_association_football", 1))


def excel_upload_dataframe_to_google_spreadsheet(auth, spreadsheet_url="", sheet_name="Sheet1", df=""):
    # Description:
    """
    Description:
        Uploads dataframe to google spreadsheet.

    Args:
        URL (str, optional): (Only in Windows)Name of Window you want to activate.
        Eg: Notepad. Defaults to "".

    Returns:
        [status]
        status (bool): Whether the function is successful or failed.
    """

    # import section
    from my_autopylot.CrashHandler import report_error
    from gspread_dataframe import set_with_dataframe
    import pandas as pd

    # Response section
    error = None
    status = False

    # Logic section

    try:
        if not auth:
            raise Exception(
                "Please call authenticate_google_spreadsheet function to get auth")

        if not spreadsheet_url:
            raise Exception("spreadsheet url cannot be empty")

        if not isinstance(df, pd.DataFrame):
            raise Exception("dataframe must be a pandas dataframe")

        sh = auth.open_by_url(url=spreadsheet_url)

        # get all the worksheets from sh
        worksheet_list = sh.worksheets()

        # check if sheet_name is already present in worksheet_list
        sheet_present = False
        for worksheet in worksheet_list:
            if worksheet.title == sheet_name:
                sheet_present = True
                break

        if sheet_present:
            # append df to existing sheet
            worksheet = sh.worksheet(sheet_name)
            row_count = worksheet.get_all_values().__len__()

            if row_count == 0:
                set_with_dataframe(worksheet, dataframe=df)
            else:
                set_with_dataframe(worksheet, dataframe=df,
                                   row=row_count + 1, include_column_header=False)

        else:
            worksheet = sh.add_worksheet(
                title=sheet_name, rows="999", cols="26")
            set_with_dataframe(worksheet, df)

    except Exception as ex:
        # check if it is a permission error
        if 'PERMISSION_DENIED' in str(ex):
            raise Exception(
                "Permission Denied. Please share the spreadsheet with client email as per Credential JSON File")
        else:
            report_error(ex)
            error = ex

    else:
        status = True
        # # If the function returns a value, it should be assigned to the data variable.
    finally:
        if error is not None:
            raise Exception(error)
        return [status]

# status, auth = authenticate_google_spreadsheet(r"C:\Users\mrmay\OneDrive\Desktop\Brainly-ML-Project\mayur-pybots-valued-door-353312-0a3451b27ef8.json")
# # status, df =excel_get_dataframe_from_google_spreadsheet(auth,"https://docs.google.com/spreadsheets/d/1CeF1NuAVLJMEBWQIT2nKceeVM9xjgxhqbNdjPBVurqw/edit#gid=0","Sheet1")

# # print(df)
# # print(status)
# # create a dummy dataframe
# import pandas as pd
# df1 = pd.DataFrame({"A":[111,222,333],"B":[4,5,6],"C":[7,8,9]})
# status = excel_upload_dataframe_to_google_spreadsheet(auth, "https://docs.google.com/spreadsheets/d/1CeF1NuAVLJMEBWQIT2nKceeVM9xjgxhqbNdjPBVurqw/edit#gid=0", "Sheet3",df1)


def excel_create_file(output_folder="", output_filename="", output_sheetname="Sheet1"):
    """
    Description:
        Creates an Excel file.
    Args:
        output_folder (str, optional): Folder where file will be created. Defaults to "".
        output_filename (str, optional): Name of file. Defaults to "".
        output_sheetname (str, optional): Name of sheet. Defaults to "Sheet1".

    Returns:
        [status]
        status (bool): Whether the function is successful or failed.
    """

    # Import section
    from my_autopylot.CrashHandler import report_error
    import os
    from pathlib import Path
    from openpyxl import Workbook

    # Response section
    error = None
    status = False

    # Logic section
    try:
        if not output_folder:
            raise Exception("Excel File name cannot be empty")

        if not output_filename:
            raise Exception("Excel File Name cannot be empty")

        if not os.path.exists(output_folder):
            output_folder = output_folder_path

        if ".xlsx" not in output_filename:
            output_filename = os.path.join(output_folder, str(
                Path(output_filename).stem) + ".xlsx")
        else:
            output_filename = os.path.join(output_folder, output_filename)

        wb = Workbook()
        ws = wb.active
        ws.title = output_sheetname

        wb.save(filename=output_filename)

    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [status]
# [status]


def valid_data(input_filepath, input_sheetname: str = "", validate_filepath=True, validate_sheetname=True):
    """
    Description:
        Validates data in excel file.

    Args:
        input_filepath (str, optional): Filepath of input file. Defaults to "".
        input_sheetname (str, optional): Sheetname of input file. Defaults to "".
        validate_filepath (bool, optional): Whether to validate filepath. Defaults to True.
        validate_sheetname (bool, optional): Whether to validate sheetname. Defaults to True.

    Returns:
        [status]
        status (bool): Whether the function is successful or failed.
    """

    import os
    from openpyxl import load_workbook
    from my_autopylot.CrashHandler import report_error

    try:
        input_filepath = str(input_filepath)
        input_sheetname = str(input_sheetname)
        if validate_filepath:
            if not ".xlsx" in input_filepath:
                raise Exception(
                    "Please provide the excel file name with .xlsx extension")
                return False
            if not os.path.exists(input_filepath):
                raise Exception(
                    "Please provide the excel file name with correct path")
                return False
            if validate_sheetname:
                wb = load_workbook(input_filepath)
                sheet_names = wb.sheetnames
                if input_sheetname not in sheet_names:
                    raise Exception(
                        "Please provide the correct sheet name")
                    print('Available Sheet Names', sheet_names)
                    return False
        return True
    except Exception as ex:
        report_error(ex)
        error = ex
        return False


def excel_to_dataframe(input_filepath="", input_sheetname="Sheet1", header=1):
    """
    Description:
        Converts excel to dataframe
    Args:
        input_filepath (str) : Complete path to the excel file.
        input_sheetname (str) : Sheet name of the excel file.
        header (int)         : Row number of the header.
    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (pandas dataframe): Dataframe of the excel file.
    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False
    data = None

    try:
        if not input_filepath:
            raise Exception("Please provide the excel path")
        if not input_sheetname:
            raise Exception("Please provide the sheet name")

        if not valid_data(input_filepath, input_sheetname):
            return [status]

        if header > 0:
            data = pd.read_excel(
                input_filepath, sheet_name=input_sheetname, header=header-1, engine='openpyxl')

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]
# df = excel_to_dataframe(r"C:\Users\mrmay\OneDrive\Desktop\MMV.xlsx")
# print(df, "df")


def excel_get_row_column_count(df):

    # Description:
    """
    Description:
        Returns the row and column count of the dataframe.


    Args:
        df (pandas dataframe): Dataframe of the excel file.

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (list): [row_count, column_count] 
    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False
    data = None

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        row, col = df.shape
        row = row + 1
        data = [row, col]

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]
# [status, data]
# data = [row, col]


def dataframe_to_excel(df, output_folder="", output_filename="", output_sheetname="Sheet1", mode='a'):  # append / overwrite
    """
    Description:
        Converts dataframe to excel
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        output_folder (str, optional): Folder path of the output file. Defaults to "".
        output_filename (str, optional): Filename of the output file. Defaults to "".
        output_sheetname (str, optional): Sheetname of the output file. Defaults to "Sheet1".
        mode (str, optional): Mode of the output file. Defaults to "a" or "x"

    Returns:
        [status]
        status (bool): Whether the function is successful or failed.
    """

    # import section
    import pandas as pd
    import os
    from pathlib import Path
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False

    try:
        if not output_folder:
            output_folder = output_folder_path
        if not output_filename:
            output_filename = "excel_file"

        if not os.path.exists(output_folder):
            output_folder = output_folder_path

        if ".xlsx" not in output_filename:
            output_filepath = os.path.join(output_folder, str(
                Path(output_filename).stem) + ".xlsx")
        else:
            output_filepath = os.path.join(output_folder, str(
                output_filename))

        if not output_sheetname:
            raise Exception("Please provide the sheet name")

        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        new_file = False
        if not os.path.exists(output_filepath):
            # excel_create_file(output_folder, output_filename)
            new_file = True

        if mode == 'a' and not new_file:
            with pd.ExcelWriter(output_filepath, mode="a", engine="openpyxl", if_sheet_exists="overlay",) as writer:
                current_df = excel_to_dataframe(
                    output_filepath, output_sheetname)[1]
                row_count = excel_get_row_column_count(current_df)[1]
                df.to_excel(writer, sheet_name=output_sheetname,
                            index=False, startrow=int(row_count[0]), header=False)
        else:
            with pd.ExcelWriter(output_filepath, engine="openpyxl",) as writer:
                df.to_excel(writer, sheet_name=output_sheetname, index=False)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status]

# if __name__ == "__main__":
#     import pandas as pd
    # df = pd.DataFrame({"A": [11, 22, 33], "B": [44, 55, 66]})
    # dataframe_to_excel(df, r"C:\Users\mrmay\OneDrive\Desktop", "MMV-2.xlsx", "Sheet1", mode='x')
    # dataframe_to_excel(df, r"C:\Users\mrmay\OneDrive\Desktop", "MMV-1.xlsx", "Sheet1")


def excel_set_single_cell(df, column_name="", cell_number=1, text=""):
    """
    Description:
        Writes the given text to the desired column/cell number for the given excel file
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        column_name (str, optional): Column name of the excel file. Defaults to "".
        cell_number (int, optional): Cell number of the excel file. Defaults to 1.
        text (str, optional): Text to be written to the excel file. Defaults to "".

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (df): Modified dataframe
    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        if not column_name:
            raise Exception("Please provide the column name")

        if not text:
            raise Exception("Please provide the text to be set")

        if cell_number < 1:
            raise Exception("Please provide the valid cell number")

        df.at[cell_number-1, column_name] = text
        data = df

    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [data]
# [status, data]
# data is a dataframe object

# df = excel_to_dataframe(r"C:\Users\mrmay\OneDrive\Desktop\MMV-1.xlsx", "Sheet1")[1]
# print(df)
# d = excel_set_single_cell(df, column_name="Salary", cell_number=3, text="Hello")[1]
# print(d)


def excel_get_single_cell(df, header=1, column_name="", cell_number=1):
    """
    Description:
        Gets the text from the desired column/cell number of the given excel file
    Args:
        df (pandas dataframe): Dataframe of the excel file.
        header (int, optional): Header of the excel file. Defaults to 0.
        column_name (str, optional): Column name of the excel file. Defaults to "".
        cell_number (int, optional): Cell number of the excel file. Defaults to 0.

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (str): Data from the desired column/cell number of the excel file.

    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        if not column_name:
            raise Exception("Please provide the column name")

        if not isinstance(column_name, list):
            column_name = [column_name]

        if cell_number < 1:
            raise Exception("Please provide the valid cell number")

        data = df.at[cell_number-header-1, column_name[0]]

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [data]
# [status, data]
# data = 'text'
# df = excel_to_dataframe(r"C:\Users\mrmay\OneDrive\Desktop\MMV-1.xlsx", "Sheet1")[1]
# print(df)
# s, d = excel_get_single_cell(df, 1, "Age", 7)
# print(d)


def excel_get_all_header_columns(df):

    # Description:
    """
    Description:
        Gives you all column header names of the given excel sheet.
    Args:
        df (pandas dataframe): Dataframe of the excel file.

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (list): List of all column header names of the excel file.

    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error
    # Response section
    error = None
    status = False
    data = None

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        data = df.columns.values.tolist()

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]
# [status, data]
# data = ['Header1', 'Header2', 'Header3']
# print(excel_get_all_header_columns(r"C:\Users\PyBots\Desktop\dummy.xlsx"))


def excel_get_all_sheet_names(input_filepath=""):
    # Description:
    """
    Description:
        Gives you all names of the sheets in the given excel sheet.

    Parameters:
        input_filepath  (str) : Path of the excel file.

    returns :
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (list): List of all sheet names of the excel file.
    """

    # import section
    from openpyxl import load_workbook
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False
    data = None

    try:
        if not input_filepath:
            raise Exception("Please provide the excel path")
        if not valid_data(input_filepath, validate_sheetname=False):
            return [status]

        wb = load_workbook(input_filepath)
        data = wb.sheetnames

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        # except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]
# [status, data]
# data = [sheet_name1, sheet_name2, sheet_name3]
# print(excel_get_all_sheet_names(r"C:\Users\PyBots\Desktop\dummy.xlsx"))


def excel_copy_range_from_sheet(input_filepath="", input_sheetname='Sheet1', start_row=1, start_col=1, end_row=1, end_col=1):

    # Description:
    """
    Description:
        Copies the specific range from the provided excel sheet and returns copied data as a list
    Args:
        input_filepath :"Full path of the excel file with double slashes"
        input_sheetname     :"Source sheet name from where contents are to be copied"
        start_col          :"Starting column number (index starts from 1) from where copying starts"
        start_row          :"Starting row number (index starts from 1) from where copying starts"
        end_col            :"Ending column number ex:4 upto where cells to be copied"
        end_row            :"Ending column number ex:5 upto where cells to be copied"

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (list): Range of cells as a list.

    """

    # import section
    from openpyxl import load_workbook
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False
    data = None
    start_row = int(start_row)
    start_col = int(start_col)
    end_row = int(end_row)
    end_col = int(end_col)

    try:

        if not input_filepath:
            raise Exception("Please provide the excel path")
        if not input_sheetname:
            raise Exception("Please provide the sheet name")

        if start_col == 0 and start_row == 0 and end_col == 0 and end_row == 0:
            raise Exception("Please provide the range to be copied.")
        if not valid_data(input_filepath, input_sheetname):
            return [False, None]

        from_wb = load_workbook(filename=input_filepath)
        try:
            fromSheet = from_wb[input_sheetname]
        except:
            fromSheet = from_wb.worksheets[0]
        rangeSelected = []

        if end_row < start_row:
            end_row = start_row

        # Loops through selected Rows
        for i in range(start_row, end_row + 1, 1):
            # Appends the row to a RowSelected list
            rowSelected = []
            for j in range(start_col, end_col+1, 1):
                rowSelected.append(
                    fromSheet.cell(row=i, column=j).value)
            # Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)

        data = rangeSelected

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]
# [status, data]
# data = [['A1', 'B1', 'C1'], ['A2', 'B2', 'C2'], ['A3', 'B3', 'C3']]
# Which can be passed to the below function as a list.


def excel_paste_range_to_sheet(input_filepath="", input_sheetname='Sheet1', start_row=1, start_col=1, copied_data=[]):

    # Description:
    """
    Description:
        Pastes the copied data in specific range of the given excel sheet.
    Args:
        input_filepath :"Full path of the excel file with double slashes"
        input_sheetname     :"Source sheet name from where contents are to be copied"
        start_col          :"Starting column number (index starts from 1) from where copying starts"
        start_row          :"Starting row number (index starts from 1) from where copying starts"
        copied_data        :"The copied data to be pasted"

    Returns:
        [status, data]
        status (bool): Whether the function is successful or failed.
        data (list): Range of cells as a list.

    """

    # import section
    from openpyxl import load_workbook
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False

    try:
        if not copied_data:
            raise Exception(
                "Copied data is empty. First copy the data using Copy Range From Sheet function.")

        if not input_filepath:
            raise Exception("Excel path is empty.")

        if start_col == 0 or start_row == 0:
            raise Exception("Please provide the range to be copied.")

        if not valid_data(input_filepath, input_sheetname):
            return [False, None]

        to_wb = load_workbook(filename=input_filepath)

        toSheet = to_wb[input_sheetname]

        # Get the number of rows and columns in the copied data
        rowCount = len(copied_data)
        colCount = len(copied_data[0])

        endRow = start_row + rowCount - 1
        endCol = start_col + colCount - 1

        countRow = 0
        for i in range(start_row, endRow+1, 1):
            countCol = 0
            for j in range(start_col, endCol+1, 1):
                toSheet.cell(
                    row=i, column=j).value = copied_data[countRow][countCol]
                countCol += 1
            countRow += 1
        to_wb.save(input_filepath)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except PermissionError:
        raise Exception(
            "Permission denied. Please close the file and try again.")
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status]
# [status]
# copied_data = [['A1', 'B1', 'C1'], ['A2', 'B2', 'C2'], ['A3', 'B3', 'C3']]
# can be obtained from the above function.

# s, d = excel_copy_range_from_sheet(r"C:\Users\mrmay\OneDrive\Desktop\MMV-1.xlsx", "Sheet1", 1, 1, 7, 2)
# print(d)
# s = excel_paste_range_to_sheet(r"C:\Users\mrmay\OneDrive\Desktop\MMV-1.xlsx", "Sheet2", 0, 0, d)


def excel_group_by_column_values_n_split(df, column_name="", output_folder="", output_filename="", show_output=False):
    """
    Description:
        This function groups the dataframe by the given column and splits the dataframe into multiple dataframes.
    Parameters:
        df : dataframe
        column_name : column name to be grouped
        output_folder : folder path to save the split dataframes
        output_filename : filename to save the split dataframes
    Returns:
        [status]
        status : True if the function is successful, False otherwise
    """

    # import section
    import pandas as pd
    from pathlib import Path
    import os
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False

    try:

        if not column_name:
            raise Exception("Please provide the column name to split.")
        if not output_folder:
            output_folder = output_folder_path
        if not output_filename:
            output_filename = "excel_split"

        grouped_df = df.groupby(column_name)

        for i in grouped_df:
            index_count = i[1].index[0] + 1
            file_path = os.path.join(
                output_folder, output_filename + "_" + str(index_count) + ".xlsx")
            file_path = Path(file_path)
            grouped_df.get_group(i[0]).to_excel(
                file_path, index=False)

        if show_output:
            os.startfile(output_folder)
        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex
        # except Exception as ex:

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status]
# [status]
# Opens the folder in which the excel files are saved.
# print(excel_group_by_column_values_n_split(
#     r"C:\Users\PyBots\Desktop\dummy.xlsx", column_name="Head 1"))


def excel_merge_all_files(input_folder_path="", output_folder="", output_filename=""):

    # Description:
    """
    Description:
        Merges all the excel files in the given folder

    Args:
        input_folder_path :"Full path of the folder with double slashes"
        output_folder      :"Full path of the folder with double slashes"
        output_filename    :"Filename to save the merged excel file"

    Returns:
        [status]
        status : True if the function is successful, False otherwise
    """

    # import section
    import os
    import pandas as pd
    import datetime
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False

    try:

        if not input_folder_path:
            raise Exception("Input folder path is empty.")

        if not output_folder:
            output_folder = output_folder_path

        if not output_filename:
            time_stamp_now = str(datetime.datetime.now().strftime("%m-%d-%Y"))
            output_filename = "excel_merged" + time_stamp_now

        filelist = [f for f in os.listdir(
            input_folder_path) if f.endswith(".xlsx")]

        all_excel_file_lst = []

        for file1 in filelist:
            file_path = os.path.join(input_folder_path, str(file1))
            file_path = str(file_path)

            all_excel_file = pd.read_excel(
                file_path, dtype=str, engine='openpyxl')
            all_excel_file_lst.append(all_excel_file)

        appended_df = pd.concat(all_excel_file_lst)

        final_path = os.path.join(
            output_folder, output_filename + ".xlsx")
        appended_df.to_excel(final_path, index=False)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except PermissionError:
        pass
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status]
# [status]


def excel_drop_columns(df, cols=""):

    # Description:
    """
    Description:
        Drops the desired column from the given excel file

    Parameters:
        df : dataframe
        cols : column names to be dropped

    Returns:
        [status, data]
        status : True if the function is successful, False otherwise
        data : dataframe with the dropped columns

    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        if not cols:
            raise Exception(
                "Please provide the column name to be dropped.")

        if not isinstance(cols, list):
            cols = [cols]

        df.drop(cols, axis=1, inplace=True)

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [status, df]
# [status, data]
# data is the dataframe after dropping the columns.


def excel_clear_sheet(df):

    # Description:
    """
    Description:
        Clears the contents of given excel files keeping header row intact

    Args:
        df : dataframe

    Returns:
        [status, data]
        status : True if the function is successful, False otherwise
        data : dataframe with the cleared contents

    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False
    data = None

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        # Clears the contents of the sheet
        df.drop(df.index, inplace=True)

        data = df

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True
    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]
# [status, data]
# data is the dataframe after clearing the sheet.


def excel_remove_duplicates(df, column_name=""):

    # Description:
    """
    Description:
        Drops the duplicates from the desired Column of the given excel file

    Args:
        df : dataframe
        column_name : column name from which duplicates are to be removed

    Returns:
        [status, data]
        status : True if the function is successful, False otherwise
        data : dataframe with the duplicates removed

    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False
    data = None
    which_one_to_keep = "first"

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        if not column_name:
            df.drop_duplicates(keep=which_one_to_keep, inplace=True)

        else:
            if not isinstance(column_name, list):
                column_name = [column_name]
            df.drop_duplicates(subset=column_name,
                               keep=which_one_to_keep, inplace=True)

        data = df
        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [status, data]
# [status, data]
# data is the dataframe after removing the duplicates.
# print(excel_remove_duplicates(
#     r"C:\Users\PyBots\Desktop\dummy.xlsx", same_file=True))


def excel_if_value_exists(df, cols="", value=""):

    # Description:
    """
    Description:
        Check if a given value exists in given excel. Returns True / False

    Args:
        df : dataframe
        cols : column name from which the value is to be checked
        value : value to be checked

    Returns:
        [status]
        status : True if the value exists, False otherwise 
    """

    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False

    try:
        if not isinstance(df, pd.DataFrame):
            raise Exception("Please provide the dataframe")

        if not value:
            raise Exception("Please provide the value to be searched")

        if not isinstance(cols, list):
            cols = [cols]

        if cols:
            if value in df[cols].values:
                status = True
        else:
            if value in df.values:
                status = True

    except Exception as ex:
        report_error(ex)
        error = ex

    finally:
        if error is not None:
            raise Exception(error)
        return [status]
# [status]
# print(excel_if_value_exists(
#     r"C:\Users\PyBots\Desktop\dummy.xlsx", value="Col1112", cols='Head 1'))


def isNaN(value=""):

    # Description:
    """
    Description:
        Returns TRUE if a given value is NaN False otherwise

    Parameters:
        value : value to be checked

    Returns:
        [status]
        status : True if the value is NaN, False otherwise

    """

    # import section
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    status = False

    try:
        if not value:
            raise Exception(
                "Value is empty. Please give a value to check.")
        import math
        status = math.isnan(float(value))

        # If the function returns a value, it should be assigned to the data variable.
        # data = value
    except Exception as ex:
        report_error(ex)
        error = ex

    finally:
        if error is not None:
            raise Exception(error)
        return [status]
# [status]


def excel_apply_format_as_table(input_filepath="", table_style="TableStyleMedium21", input_sheetname='Sheet1'):
    '''
    Description:
        Applies table format to the used range of the given excel.
        Just it takes an path and converts it to table here you can change the table style below.
        if you want to change the table style just change the styles by refering excel

    Args:
        input_filepath : path of the excel file
        table_style : table style to be applied
        input_sheetname : sheet name of the excel file

    Returns:
        [status]
        status : True if the function is successful, False otherwise
     '''

    # Import section
    from my_autopylot.CrashHandler import report_error
    import win32com.client

    # Response section
    error = None
    status = False

    # Logic section
    try:
        if not input_filepath:
            raise Exception("Excel File name cannot be empty")

        if not input_sheetname:
            raise Exception("Sheet name cannot be empty")

        if not isinstance(input_sheetname, list):
            table_style = [table_style]

        if not isinstance(input_filepath, list):
            input_filepath = [input_filepath]

        if not isinstance(input_sheetname, list):
            input_sheetname = [input_sheetname]

        for excel_file_path_i, sheet_name_i, table_style_i in zip(input_filepath, input_sheetname, table_style):
            excel_instance = win32com.client.gencache.EnsureDispatch(
                "Excel.Application")
            excel_instance.Visible = False
            excel_instance.DisplayAlerts = False

            exc_workbook = excel_instance.Workbooks.Open(
                Filename=excel_file_path_i.replace("/", "\\"))  # .Sheets.Item[sheet_name]
            try:
                exc_workbook.Worksheets(sheet_name_i).Select()
                excel_instance.ActiveSheet.UsedRange.Select()
                excel_instance.Selection.Columns.AutoFit()
                excel_instance.ActiveSheet.ListObjects.Add().TableStyle = table_style_i
                exc_workbook.Close(SaveChanges=1)
                excel_instance.Quit()
            except:
                exc_workbook.Close()
                excel_instance.Quit()
                raise Exception("Given Excel already has a table")
    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [status]
# [status]
# print(excel_apply_format_as_table(
#     r"C:\Users\PyBots\Desktop\dummy.xlsx", table_style="TableStyleMedium21", input_sheetname='Sheet2'))


def excel_apply_template_format(input_filepath='', input_sheetname='Sheet1', input_template_filepath='', input_template_sheetname="Sheet1", same_file=True, output_folder="", output_filename=""):
    """
    Description:
        Converts given excel to Template Excel
        This function uses pandas and just write the required columns to new excel.
        if you don't know columns, just pass the excel file which have the columns you want it automatically makes own list and remove other columns.

    Args:
        input_filepath : path of the excel file
        input_sheetname : sheet name of the excel file
        input_template_filepath : path of the template excel file
        input_template_sheetname : sheet name of the template excel file
        same_file : if True, then the output excel file will be same as the input excel file.
        output_folder : folder path where the output excel file will be saved.
        output_filename : name of the output excel file.

    Returns:
        [status]
        status : True if the function is successful, False otherwise

    """
    # Import section
    from my_autopylot.CrashHandler import report_error
    import pandas as pd
    import datetime

    # Response section
    error = None
    status = False

    # Logic section
    try:
        if not input_filepath:
            raise Exception("Raw Excel File name cannot be empty")

        if not input_sheetname:
            raise Exception("Raw Excel Sheet name cannot be empty")

        if not input_template_filepath:
            raise Exception("Template Excel File name cannot be empty")

        if not output_folder:
            output_folder = output_folder_path

        if not output_filename:
            time_stamp_now = str(datetime.datetime.now().strftime("%m-%d-%Y"))
            output_filename = "excel_new_" + time_stamp_now

        cols = excel_get_all_header_columns(
            input_template_filepath, input_template_sheetname)

        df = pd.read_excel(input_filepath,
                           sheet_name=input_sheetname, usecols=cols)

        if not same_file:
            new_file_path = os.path.join(
                output_folder, output_filename + ".xlsx")
            df.to_excel(new_file_path, index=False)
        else:
            df.to_excel(input_filepath, index=False)

    except Exception as ex:
        report_error(ex)
        error = ex

    else:
        status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [status]

# write a function to create a dataframe from a list of lists


def df_from_list(list_of_lists, column_names=None):
    """
    Description:
        Creates a dataframe from a list of lists

    Args:
        list_of_lists : list of lists
        column_names : list of column names

    Returns:
        [data]
        data : dataframe object
    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None
    # Logic section
    try:
        if not isinstance(list_of_lists, list):
            raise Exception("Please pass input as list of lists")

        if isinstance(list_of_lists, list):
            if column_names == None:
                data = pd.DataFrame(list_of_lists)
            else:
                data = pd.DataFrame(list_of_lists, columns=column_names)

    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]
# [data]
# print(df_from_list([[1,2,3],[4,5,6],[7,8,9]],column_names=['a','b','c']))
# s, df1=df_from_list([[1,2,3],[4,5,6],[7,8,9]])
# print(df1, type(df1))
# print(df_from_list([[1,2,3],[4,5,6],['a','b','c']]))

# write a function to create a dataframe from a string


def df_from_string(df_string: str, word_delimeter=" ", line_delimeter="\n", column_names=None):
    """
    Description:
        Creates a dataframe from a string
    Args:
        df_string : string
        word_delimeter : word delimeter
        line_delimeter : line delimeter
        column_names : list of column names

    Returns:
        [data]
        data : dataframe object
    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None
    # Logic section
    try:
        if not df_string:
            raise Exception("Please pass input as string")

        if not isinstance(df_string, str):
            df_string = str(df_string)

        if column_names == None:
            data = pd.DataFrame([x.split(word_delimeter)
                                for x in df_string.split(line_delimeter)])
        elif isinstance(column_names, list):
            data = pd.DataFrame([x.split(word_delimeter) for x in df_string.split(
                line_delimeter)], columns=column_names)
    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]
# [data]
# print(df_from_string('1 2 3\n4 5 6\n7 8 9', word_delimeter=" ", line_delimeter="\n"))
# print(df_from_string('1\t2\t3\r\n4\t5\t6\r\n7\t8\t9', word_delimeter="\t", line_delimeter="\r\n",column_names=['a','b','c']))


# function to extract sub dataframe from a dataframe using rows and column number
def df_extract_sub_df(df, row_start: int, row_end: int, column_start: int, column_end: int):
    """
    Description:
        Extracts a sub dataframe from a dataframe
    Args:
        df : dataframe
        row_start : start row number
        row_end : end row number
        column_start : start column number
        column_end : end column number

    Returns:
        [data]
        data : dataframe object
    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None
    # Logic section
    try:
        if df.empty:
            raise Exception("Dataframe cannot be empty")

        if isinstance(df, pd.DataFrame):
            data = df.iloc[row_start:row_end, column_start:column_end]

    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]
# [status, data]
# df_main = df_from_list([[1,2,3,4,5],[6,7,8,9,10],[11,12,13,14,15],[16,17,18,19,20]],column_names=['a','b','c','d','e'])
# print(df_main,"\n")
# df_sub = df_extract_sub_df(df_main,1,2,2,3)
# print(df_sub)


def set_value_in_df(df, row_number: int, column_number: int, value):
    """
    Description:
        Sets a value in a dataframe

    Args:
        df : dataframe
        row_number : row number
        column_number : column number
        value : value to be set

    Returns:
        [data]
        data : dataframe object

    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None
    # Logic section
    try:
        if df.empty:
            raise Exception("Dataframe cannot be empty")

        # print(type(df))
        if isinstance(df, pd.DataFrame):
            if row_number < 1 or column_number < 1:
                raise Exception(
                    "Row and column number should be greater than 0")

            if row_number > df.shape[0] or column_number > df.shape[1]:
                raise Exception(
                    "Row and column number should be less than or equal to dataframe shape")

            df.iloc[row_number-1, column_number-1] = value

    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [df]
# [status, data]
# df1 = df_from_list([[1,2,3,4,5],[6,7,8,9,10],[11,12,13,14,15],[16,17,18,19,20]],column_names=['a','b','c','d','e'])
# print(df1)
# print(type(df1))
# print(set_value_in_df(df1,0,0,100))


def get_value_in_df(df, row_number: int, column_number: int):
    """
    Description:
        Gets a value from a dataframe

    Parameters:
        df : dataframe
        row_number : row number
        column_number : column number

    Returns:
        [data]
        data : value from dataframe

    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None

    # Logic section
    try:
        if df.empty:
            raise Exception("Dataframe cannot be empty")

        if isinstance(df, pd.DataFrame):
            if row_number < 1 or column_number < 1:
                raise Exception(
                    "Row and column number should be greater than 0")

            if row_number > df.shape[0] or column_number > df.shape[1]:
                raise Exception(
                    "Row and column number should be less than or equal to dataframe shape")

            data = df.iloc[row_number-1, column_number-1]

    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]
# [status, data]

# create a function to merge all sheets of an excel


def excel_concat_all_sheets_of_given_excel(excel_file_path, sheet_names_as_list=None):
    """
    Description:
        Concatenates all sheets of an excel file

    Args:
        excel_file_path : excel file path

    Returns:
        [data]
        data : dataframe object
    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None

    # Logic section
    try:
        if not excel_file_path:
            raise Exception("Please pass excel file path")

        if isinstance(excel_file_path, str):
            if sheet_names_as_list == None:
                data = pd.read_excel(excel_file_path, sheet_name=None)
            elif isinstance(sheet_names_as_list, list):
                data = pd.read_excel(
                    excel_file_path, sheet_name=sheet_names_as_list)
            data = pd.concat(data, ignore_index=True)
        else:
            raise Exception("Please pass sheet names as list")

    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]

# if __name__ == "__main__":
#     # data= excel_concat_all_sheets_of_given_excel(r"C:\Users\mrmay\OneDrive\Desktop\MMV.xlsx")
#     data= excel_concat_all_sheets_of_given_excel(r"C:\Users\mrmay\OneDrive\Desktop\MMV.xlsx",sheet_names_as_list=["Sheet1","Sheet3"])
#     print(data[0])

# write a function to drop a range of rows from dataframe


def df_drop_rows(df, row_start: int, row_end: int):
    """
    Description:
        Drops a range of rows from a dataframe including the row_start and row_end rows.

    Args:
        df : dataframe
        row_start : start row number
        row_end : end row number

    Returns:
        [data]
        data : dataframe object
    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None
    # Logic section
    try:
        if df.empty:
            raise Exception("Dataframe cannot be empty")

        if isinstance(df, pd.DataFrame):
            # -1 because index starts from 0
            data = df.drop(df.index[row_start-1:row_end])

    except Exception as ex:
        # report_error(ex)
        error = ex
        print(ex)

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]

# if __name__ == "__main__":
#     status, df = excel_to_dataframe(r"C:\Users\mrmay\OneDrive\Desktop\MMV.xlsx",input_sheetname="Sheet3")
#     print(df)
#     data= df_drop_rows(df,row_start=15,row_end=17)
#     print(data[0])

# write a function to convert dataframe column or list of columns to string / int / float /date as per user choice


def df_convert_column_to_type(df, column_name: str, column_type: str):
    """
    Description:
        Converts a column type of a dataframe to a given type
        Column type doesn't persist after writing to excel

    Args:
        df : dataframe

        column_name : Single column name or list of column names

        column_type : column type to be converted to like string, int, float, date, boolean, complex, bytes, etc.

    Returns:
        [data]
        data : The modified dataframe object
    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None
    # Logic section
    try:
        if df.empty:
            raise Exception("Dataframe cannot be empty")

        if isinstance(df, pd.DataFrame):
            if str(column_type).lower() in ["string", "str"]:
                df[column_name] = df[column_name].astype('str')
            elif str(column_type).lower() in ["int", "integer"]:
                df[column_name] = df[column_name].astype('int64')
            elif str(column_type).lower() in ["float", "double"]:
                df[column_name] = df[column_name].astype('float64')
            elif str(column_type).lower() in ["date"]:
                df[column_name] = pd.to_datetime(df[column_name], unit='s')
            elif str(column_type).lower() in ["boolean", "bool"]:
                df[column_name] = df[column_name].astype('bool')
            elif str(column_type).lower() in ["complex"]:
                df[column_name] = df[column_name].astype('complex128')
            elif str(column_type).lower() in ["bytes"]:
                df[column_name] = df[column_name].astype('bytes')
            else:
                try:
                    df[column_name] = df[column_name].astype(column_type)
                except:
                    raise Exception("Please pass correct column type")

            data = df
    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]

# if __name__ == "__main__":
#     # print(excel_create_file(r"C:\Users\mrmay\OneDrive\Desktop","MMV-9.PDF","Sheet1"))
#     status, df = excel_to_dataframe(r"C:\Users\mrmay\OneDrive\Desktop\MMV.xlsx",input_sheetname="Sheet3")
#     print(df)
#     print(df.dtypes)
#     data= df_convert_column_to_type(df,column_name="Age",column_type="double")
#     print(data[0])
#     dataframe_to_excel(data[0],r"C:\Users\mrmay\OneDrive\Desktop", "MMV.xlsx",output_sheetname="Sheet3")

# write a function to perform vlookup on 2 dataframes


def df_vlookup(df1, df2, column_name: str, how: str = "left"):
    """
    Description:
        Performs vlookup on 2 dataframes

    Args:
        df1 : dataframe
        df2 : dataframe
        column_name : column name to perform vlookup on
        how : how to perform vlookup like inner, left, right, outer

    Returns:
        [data]
        data : The modified dataframe object
    """
    # import section
    import pandas as pd
    from my_autopylot.CrashHandler import report_error

    # Response section
    error = None
    # status = False
    data = None
    # Logic section
    try:
        if df1.empty or df2.empty:
            raise Exception("Dataframe cannot be empty")

        if column_name not in df1.columns or column_name not in df2.columns:
            raise Exception("Please pass correct column name")

        if isinstance(df1, pd.DataFrame) and isinstance(df2, pd.DataFrame):
            data = pd.merge(df1, df2, on=column_name, how=how)
    except Exception as ex:
        report_error(ex)
        error = ex

    # else:
    #     status = True

    finally:
        if error is not None:
            raise Exception(error)
        return [data]

# if __name__ == "__main__":
#     s,df1 = excel_to_dataframe(r"C:\Users\mrmay\OneDrive\Desktop\MMV-1.xlsx",input_sheetname="Sheet1")
#     s,df2 = excel_to_dataframe(r"C:\Users\mrmay\OneDrive\Desktop\MMV-2.xlsx",input_sheetname="Sheet1")

#     df = df_vlookup(df1, df2, column_name="Age", how="outer")
#     print(df1)
#     print(df2)
#     print(df[0])
